from abc import abstractmethod
from tarfile import LNKTYPE
from types import DynamicClassAttribute
from openpyxl.descriptors.base import _convert
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import random
import platform
import sys
import os
import json
import pickle
import click
import openpyxl
from openpyxl import load_workbook
import random
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import WebDriverException, TimeoutException, MoveTargetOutOfBoundsException
from selenium.common.exceptions import TimeoutException
from pathlib import Path
from webdriver_manager import driver
from webdriver_manager.chrome import ChromeDriverManager
from termcolor import colored
from colorama import init
init()
wb = load_workbook('a.xlsx')
sh = wb.active
maxrow = sh.max_row
wb = openpyxl.load_workbook('a.xlsx')
sheet = wb['Sheet1']
a = random.randint(1,maxrow)
tensp = sheet.cell(row=a, column=1)
linksp = sheet.cell(row=a, column=2)
d = tensp.value 

z = 1
j = 1
p = random.randint(1,3)
BASE_PATH = Path('__file__').resolve().parent
DRIVERS_PATH = BASE_PATH / "drivers"
os_version = platform.system()
ssss =open("proxy.txt",'r')
asss = ssss.readline()
ssss.close()

class Linkedin:
    def __init__(self, port=None):
        self.port = port

    def create_driver(self, headless):
        options = Options()
        if self.port:
            options.add_experimental_option(
                "debuggerAddress", f"localhost:{self.port}")
        else:
            ua = UserAgent()
            assd = ua.random
            user_agent = ua.random
            options.add_argument(f'user-agent={user_agent}')
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option(
                "excludeSwitches", ["enable-automation"])
            
          
            options.add_argument('--log-level=3')
            prefs = {"profile.default_content_setting_values.notifications": 2}
            options.add_experimental_option("prefs", prefs)
            options.headless = headless


        if os_version == "Linux":
            executable = os.path.join(DRIVERS_PATH, "chromedriver_linux")
        elif os_version == "Darwin":
            executable = os.path.join(DRIVERS_PATH, "chromedriver_mac")
        elif os_version == "Windows":
            executable = os.path.join(DRIVERS_PATH, "chromedriver_windows.exe")
        else:
            driver = webdriver.Chrome(
                ChromeDriverManager().install(), options=options)
            return driver
        driver = webdriver.Chrome(executable_path=executable, options=options)
        return driver
    
    def  khoitao(self , driver , alp):
        try:
            
            driver.get('https://www.shopee.vn')
            time.sleep(random.randint(3,7))
            driver.find_element_by_xpath('//*[@id="modal"]/div/div/div[2]/div').click()
        except TimeoutException:
          driver.close()
    def tukhoa(self,driver,b):
        try:
            
            time.sleep(random.randint(5,10))
            driver.execute_script('window.scrollTo(0,1000)')
            driver.find_element_by_xpath("//input").send_keys(b)
            time.sleep(random.randint(1,4))
            driver.find_element_by_xpath("//input").send_keys(Keys.ENTER)
        except TimeoutException:
          driver.close()
    
        
    def tuongtacsanpham1(self , driver):
         iio =1
         while (iio<= 4):
                c123 = random.randint(1000,3000)
                ww = 'window.scrollTo(0,{})'.format(c123)
                driver.execute_script(ww)
                time.sleep(random.randint(5,10))
                driver.execute_script(ww)
                time.sleep(random.randint(5,10))
                driver.execute_script(ww)
                time.sleep(random.randint(5,10))
                iio = iio + 1
                
                    
         profiles = BeautifulSoup(driver.page_source)
         link = profiles.find_all('div', class_='_10Wbs- _5SSWfi UjjMrh')
                        
                
                
         aa1 = '"{}"'.format(link)
                    

         bb1 = aa1.replace('</div>',"-")
         cc1 = bb1.replace('<div class="_10Wbs- _5SSWfi UjjMrh">', '-ta-').replace('[-ta-','').replace('"','').replace('-]"','')


         def Convert(string):
                li = list(string.split('-, -ta-'))
                return li
         asd  = Convert(cc1) 
         return (asd)
                    
                    
                   
                    

     

    def clicksanpham(self, driver, a12):
                a18 =  a12 + 1
                c12 = '//*[@id="main"]/div/div[3]/div/div[2]/div[2]/div[2]/div[{}]'.format(a18)
                driver.find_element_by_xpath(c12).click()
                

    def  chuyentrang(self , driver):
                
            
            try:
                    
                 next = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                    (By.XPATH, "//button[contains(@class, 'shopee-icon-button shopee-icon-button--right')]")))
                 driver.execute_script( "return arguments[0].scrollIntoView(true);", next)
                 time.sleep(random.randint(1,5))
                 ActionChains(driver).move_to_element(next).click().perform()
                        
            except TimeoutException:
                driver.close()
       
    def tuongtacsanpham(self,driver, assc):
       
        try:
            
            time.sleep(random.randint(5,10))
            driver.execute_script('window.scrollTo(0,2000)')
           
            submit_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class = '_3-_YTZ']")))
            ActionChains(driver).move_to_element(submit_button).click().perform()
            i = random.randint(4,5)
            while (i <= 8):
                time.sleep(random.randint(5,10))
                ads= WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class = '_2cmzNg']")))
                ActionChains(driver).move_to_element(ads).click().perform()
                i = i + 1
            
            driver.refresh()
            iio =1
            while (iio<= 4): 
                    c123 = random.randint(1000,3000)
                    ww = 'window.scrollTo(0,{})'.format(c123)
                    driver.execute_script(ww)
                    time.sleep(random.randint(5,10))
                    
                    iio = iio + 1
            g  =  random.randint(2,4)
            try:
                while(g <= 5):
                    vcl1 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//div[contains(@class, 'product-rating-overview__filter')]")))
                    time.sleep(random.randint(5,10)) 
                    driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl1)
                    time.sleep(random.randint(5,10))
                    ActionChains(driver).move_to_element(vcl1).click().perform()
                    driver.execute_script('window.scrollTo(0,2000)')
                    g = g + 1 
                    time.sleep(random.randint(5,10))
            except TimeoutException:
                pass  
            try:
                cc=  random.randint(1,5)
                while(cc <= 5):
                    vcl12 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//button[contains(@class, 'shopee-icon-button shopee-icon-button--right')]")))
                    time.sleep(random.randint(10,15)) 
                    driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl12)
                    time.sleep(random.randint(10,15))
                    ActionChains(driver).move_to_element(vcl12).click().perform()
                    cc = cc + 1
            except TimeoutException:
                pass
        except TimeoutException:
             driver.close()
    def sanphamtuongtu(self,driver,sda):
         try:
         
            time.sleep(random.randint(5,10))
            vcl124 = WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                                                (By.XPATH, "//div[contains(@class, 'product-recommend-items__item-wrapper')]")))
            
            driver.execute_script( "return arguments[0].scrollIntoView(true);", vcl124)
            
            ActionChains(driver).move_to_element(vcl124).click().perform()
         except TimeoutException:
            pass
    def delete_cache(self , driver ):
        driver.execute_script("window.open('');")
        time.sleep(2)
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(2)
        driver.get('chrome://settings/clearBrowserData') # for old chromedriver versions use cleardriverData
        time.sleep(2)
        actions = ActionChains(driver) 
        actions.send_keys(Keys.TAB * 3 + Keys.DOWN * 3) # send right combination
        actions.perform()
        time.sleep(2)
        actions = ActionChains(driver) 
        actions.send_keys(Keys.TAB * 4 + Keys.ENTER) # confirm
        actions.perform()
        time.sleep(5) # wait some time to finish
        driver.close() # close this tab
        driver.switch_to.window(driver.window_handles[0]) # switch back
        
        driver.delete_all_cookies()


@click.command()
@click.option("-h", "--headless", is_flag=True, help="Hide the borwser's window.")

def main(headless):
    try:
         ln = Linkedin()
         driver12 = ln.create_driver('a')
         ln.khoitao(driver12, 'a')
         ln.tukhoa(driver12,d)
         
         
         asd1 = ln.tuongtacsanpham1(driver12)
                  
         try:
             w1 = asd1.index(linksp.value)
             ln.clicksanpham(driver12,w1)
         except:
             ln.chuyentrang(driver12)
             asd12 = ln.tuongtacsanpham1(driver12)
             try:
                w12= asd12.index(linksp.value)
                ln.clicksanpham(driver12,w12)
             except:
                 ln.chuyentrang(driver12)
                 asd13 = ln.tuongtacsanpham1(driver12)
                 try:
                    w123= asd13.index(linksp.value)
                    ln.clicksanpham(driver12,w123)
                 except:

                     ln.chuyentrang(driver12)
                     asd1234 = ln.tuongtacsanpham1(driver12)
                     try:
                        w1235= asd1234.index(linksp.value)
                        ln.clicksanpham(driver12,w1235)
                     except:
                         pass
                                



                

   

         iii = random.randint(2,3)
         while (iii <= 6):
             ln.tuongtacsanpham(driver12,'d')
             time.sleep(random.randint(5,10))
             ln.sanphamtuongtu(driver12,'q')
             iii = iii + 1
         ln.delete_cache(driver12)
    except TimeoutException:
        
        ln.close()
      
if __name__ == "__main__":
    main()  


